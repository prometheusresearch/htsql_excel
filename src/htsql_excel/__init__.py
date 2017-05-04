#
# Copyright (c) 2016, Prometheus Research, LLC
#

try:
    from cStringIO import StringIO
except ImportError:
    from StringIO import StringIO

import math

import openpyxl
import xlwt

from htsql.core.adapter import Adapter, adapt, adapt_many, call
from htsql.core.addon import Addon
from htsql.core.cmd.summon import SummonFormat
from htsql.core.fmt.accept import Accept
from htsql.core.fmt.format import Format
from htsql.core.fmt.emit import EmitHeaders, Emit
from htsql.core.domain import Domain, BooleanDomain, NumberDomain, \
    FloatDomain, DecimalDomain, TextDomain, EnumDomain, DateDomain, \
    TimeDomain, DateTimeDomain, ListDomain, RecordDomain, UntypedDomain, \
    VoidDomain, OpaqueDomain, Profile
from htsql.core.util import listof


XLS_MIME_TYPE = 'application/vnd.ms-excel'
XLSX_MIME_TYPE = \
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class ExcelAddon(Addon):
    name = 'htsql_excel'
    hint = 'Basic support for Excel files'


class ToExcel(Adapter):
    adapt(Domain)

    def __init__(self, domain, profiles):
        assert isinstance(domain, Domain)
        assert isinstance(profiles, listof(Profile)) and len(profiles) > 0
        self.domain = domain
        self.profiles = profiles
        self.width = 1

    def __call__(self):
        return self

    def headers(self):
        return [self.profiles[-1].header]

    def cells(self, value):
        if value is None:
            yield [None]
        else:
            yield [self.domain.dump(value)]


class VoidToExcel(ToExcel):
    adapt(VoidDomain)

    def __init__(self, domain, profiles):
        super(VoidToExcel, self).__init__(domain, profiles)
        self.width = 0

    def headers(self):
        return []

    def cells(self, value):
        yield []


class RecordToExcel(ToExcel):
    adapt(RecordDomain)

    def __init__(self, domain, profiles):
        super(RecordToExcel, self).__init__(domain, profiles)
        self.fields_to_excel = [
            to_excel(field.domain, profiles + [field])
            for field in domain.fields
        ]
        self.width = 0
        for field_to_excel in self.fields_to_excel:
            self.width += field_to_excel.width

    def headers(self):
        row = []
        for field_to_excel in self.fields_to_excel:
            row.extend(field_to_excel.headers())
        return row

    def cells(self, value):
        if not self.width:
            return
        if value is None:
            yield [None] * self.width
        else:
            streams = [
                (field_to_excel.cells(item), field_to_excel.width)
                for item, field_to_excel in zip(value, self.fields_to_excel)
            ]
            is_done = False
            while not is_done:
                is_done = True
                row = []
                for stream, width in streams:
                    subrow = next(stream, None)
                    if subrow is None:
                        subrow = [None] * width
                    else:
                        is_done = False
                    row.extend(subrow)
                if not is_done:
                    yield row


class ListToExcel(ToExcel):
    adapt(ListDomain)

    def __init__(self, domain, profiles):
        super(ListToExcel, self).__init__(domain, profiles)
        self.item_to_excel = to_excel(domain.item_domain, profiles)
        self.width = self.item_to_excel.width

    def headers(self):
        return self.item_to_excel.headers()

    def cells(self, value):
        if not self.width:
            return
        if value is not None:
            item_to_cells = self.item_to_excel.cells
            for item in value:
                for row in item_to_cells(item):
                    yield row


class SimpleToExcel(ToExcel):
    adapt_many(
        BooleanDomain,
        NumberDomain,
        UntypedDomain,
        TextDomain,
        EnumDomain,
        DateDomain,
        TimeDomain,
        DateTimeDomain,
    )

    def cells(self, value):
        yield [value]


class FloatToExcel(ToExcel):
    adapt(FloatDomain)

    def cells(self, value):
        if value is None or math.isinf(value) or math.isnan(value):
            yield [None]
        else:
            yield [value]


class DecimalToExcel(ToExcel):
    adapt(DecimalDomain)

    def cells(self, value):
        if value is None or not value.is_finite():
            yield [None]
        else:
            yield [value]


class OpaqueToExcel(ToExcel):
    adapt(OpaqueDomain)

    def cells(self, value):
        if value is None:
            yield [None]
            return
        if not isinstance(value, unicode):
            try:
                value = str(value).decode('utf-8')
            except UnicodeDecodeError:
                value = unicode(repr(value))
        yield [value]


to_excel = ToExcel.__invoke__  # pylint: disable=invalid-name


def make_name(meta):
    filename = None
    if meta.header:
        filename = meta.header.encode('utf-8')
    if not filename:
        filename = 'data'
    filename = filename.replace('\\', '\\\\').replace('"', '\\"')
    return filename


class EmitExcelHeaders(EmitHeaders):
    def __call__(self):
        yield (
            'Content-Type',
            self.content_type,
        )
        yield (
            'Content-Disposition',
            'attachment; filename="%s.%s"' % (
                make_name(self.meta),
                self.file_extension,
            ),
        )


class EmitExcel(Emit):
    def __call__(self):
        product = to_excel(self.meta.domain, [self.meta])
        output = StringIO()
        self.render(output, product)
        yield output.getvalue()


class ExcelFormat(Format):
    pass


class XLSFormat(ExcelFormat):
    pass


class SummonXLS(SummonFormat):
    call('xls')
    format = XLSFormat


class AcceptXLS(Accept):
    call(XLS_MIME_TYPE)
    format = XLSFormat


class EmitXLSHeaders(EmitExcelHeaders):
    adapt(XLSFormat)

    content_type = XLS_MIME_TYPE
    file_extension = 'xls'


class EmitXLS(EmitExcel):
    adapt(XLSFormat)

    def render(self, stream, product):
        # Build the file
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet(make_name(self.meta))

        # Headers
        style = xlwt.XFStyle()
        style.font = xlwt.Font()
        style.font.bold = True
        for idx, header in enumerate(product.headers()):
            sheet.write(0, idx, header, style)

        # Data
        for row_idx, row in enumerate(product.cells(self.data)):
            for cell_idx, cell in enumerate(row):
                sheet.write((row_idx + 1), cell_idx, cell)

        workbook.save(stream)


class XLSXFormat(ExcelFormat):
    pass


class SummonXLSX(SummonFormat):
    call('xlsx')
    format = XLSXFormat


class AcceptXLSX(Accept):
    call(XLSX_MIME_TYPE)
    format = XLSXFormat


class EmitXLSXHeaders(EmitExcelHeaders):
    adapt(XLSXFormat)

    content_type = XLSX_MIME_TYPE
    file_extension = 'xlsx'


class EmitXLSX(EmitExcel):
    adapt(XLSXFormat)

    def render(self, stream, product):
        # Build the file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = make_name(self.meta)

        # Headers
        font = openpyxl.styles.Font(bold=True)
        for idx, header in enumerate(product.headers()):
            cell = sheet.cell(row=1, column=(idx + 1))
            cell.value = header
            cell.font = font

        # Data
        for row in product.cells(self.data):
            sheet.append(row)

        workbook.save(stream)

